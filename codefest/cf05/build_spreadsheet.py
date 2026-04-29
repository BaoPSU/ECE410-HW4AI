"""
CF05 Systolic Array — Excel spreadsheet generator.
Produces systolic_trace.xlsx with four sheets:
  1. Setup & Verification
  2. PE Weights
  3. Cycle-by-Cycle Trace
  4. Counts (MACs, Reuse, Off-chip)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── colour palette ───────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
PALE_BLUE   = "DEEAF1"
YELLOW      = "FFE699"
GREEN       = "E2EFDA"
DARK_GREEN  = "375623"
ORANGE      = "FCE4D6"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"
DARK_GREY   = "595959"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def style(ws, cell_ref, value=None, bold=False, bg=None, fg="000000",
          size=11, align="center", italic=False, border=True, formula=None):
    c = ws[cell_ref]
    if formula is not None:
        c.value = formula
    elif value is not None:
        c.value = value
    c.font = Font(bold=bold, color=fg, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=True
    )
    if border:
        c.border = thin_border()
    return c

def header(ws, cell_ref, value, bg=DARK_BLUE, fg=WHITE, size=12):
    style(ws, cell_ref, value=value, bold=True, bg=bg, fg=fg,
          size=size, align="center")

def sub_header(ws, cell_ref, value, bg=MED_BLUE, fg=WHITE):
    style(ws, cell_ref, value=value, bold=True, bg=bg, fg=fg,
          size=11, align="center")

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def merge_header(ws, cell_range, value, bg=DARK_BLUE, fg=WHITE, size=12):
    ws.merge_cells(cell_range)
    top_left = cell_range.split(":")[0]
    style(ws, top_left, value=value, bold=True, bg=bg, fg=fg,
          size=size, align="center")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Setup & Verification
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1 - Setup & Verification"
ws1.sheet_view.showGridLines = False

for col, w in enumerate([3, 14, 10, 10, 3, 14, 10, 10, 3, 14, 10, 10, 3], 1):
    set_col_width(ws1, col, w)
for row in range(1, 60):
    set_row_height(ws1, row, 20)

# Title
ws1.merge_cells("A1:M1")
style(ws1, "A1", value="CF05 — Weight-Stationary 2×2 Systolic Array",
      bold=True, bg=DARK_BLUE, fg=WHITE, size=14, align="center", border=False)
set_row_height(ws1, 1, 30)

# ── Matrix A ──────────────────────────────────────────────────────────────────
merge_header(ws1, "A3:D3", "Matrix A  (2×2)", bg=MED_BLUE)
sub_header(ws1, "B4", "Col 0"); sub_header(ws1, "C4", "Col 1")
sub_header(ws1, "A4", "", bg=MED_BLUE)
style(ws1, "A5", "Row 0", bold=True, bg=LIGHT_BLUE, align="center")
style(ws1, "A6", "Row 1", bold=True, bg=LIGHT_BLUE, align="center")
for cell, val in [("B5", 1), ("C5", 2), ("B6", 3), ("C6", 4)]:
    style(ws1, cell, value=val, bg=WHITE, size=12)

# blank col D filler
ws1["D4"].value = ""
ws1["D5"].value = ""
ws1["D6"].value = ""

# ── Matrix B ──────────────────────────────────────────────────────────────────
merge_header(ws1, "F3:I3", "Matrix B  (2×2, weights)", bg=MED_BLUE)
sub_header(ws1, "G4", "Col 0"); sub_header(ws1, "H4", "Col 1")
sub_header(ws1, "F4", "", bg=MED_BLUE)
style(ws1, "F5", "Row 0", bold=True, bg=LIGHT_BLUE, align="center")
style(ws1, "F6", "Row 1", bold=True, bg=LIGHT_BLUE, align="center")
for cell, val in [("G5", 5), ("H5", 6), ("G6", 7), ("H6", 8)]:
    style(ws1, cell, value=val, bg=WHITE, size=12)

# ── Expected C ────────────────────────────────────────────────────────────────
merge_header(ws1, "K3:M3", "Expected  C = A×B", bg=DARK_GREEN, fg=WHITE)
sub_header(ws1, "L4", "Col 0", bg=DARK_GREEN, fg=WHITE)
sub_header(ws1, "M4", "Col 1", bg=DARK_GREEN, fg=WHITE)
sub_header(ws1, "K4", "", bg=DARK_GREEN, fg=WHITE)
style(ws1, "K5", "Row 0", bold=True, bg=GREEN, align="center")
style(ws1, "K6", "Row 1", bold=True, bg=GREEN, align="center")
for cell, val in [("L5", 19), ("M5", 22), ("L6", 43), ("M6", 50)]:
    style(ws1, cell, value=val, bg=GREEN, size=12, bold=True)

# ── Verification section ──────────────────────────────────────────────────────
merge_header(ws1, "A9:M9", "Step-by-Step Verification:  C[i][j] = Σ A[i][k] × B[k][j]",
             bg=MED_BLUE)
set_row_height(ws1, 9, 24)

# Column headers
for col, label in enumerate(
    ["Element", "Formula", "A[i][0]", "×", "B[0][j]", "+",
     "A[i][1]", "×", "B[1][j]", "=", "Sum", "Expected", "Match?"], 1):
    sub_header(ws1, f"{get_column_letter(col)}10", label, bg=LIGHT_BLUE, fg="000000")
set_row_height(ws1, 10, 22)

rows = [
    ("C[0][0]", "A[0][0]×B[0][0] + A[0][1]×B[1][0]",
     "=B5", "×", "=G5", "+", "=C5", "×", "=G6",
     "=", "=C11*E11+G11*I11", 19),
    ("C[0][1]", "A[0][0]×B[0][1] + A[0][1]×B[1][1]",
     "=B5", "×", "=H5", "+", "=C5", "×", "=H6",
     "=", "=C12*E12+G12*I12", 22),
    ("C[1][0]", "A[1][0]×B[0][0] + A[1][1]×B[1][0]",
     "=B6", "×", "=G5", "+", "=C6", "×", "=G6",
     "=", "=C13*E13+G13*I13", 43),
    ("C[1][1]", "A[1][0]×B[0][1] + A[1][1]×B[1][1]",
     "=B6", "×", "=H5", "+", "=C6", "×", "=H6",
     "=", "=C14*E14+G14*I14", 50),
]

for r_off, (elem, formula, a0, x1, b0, plus, a1, x2, b1, eq, sumf, exp) \
        in enumerate(rows, 11):
    row = r_off
    bg = PALE_BLUE if r_off % 2 == 1 else WHITE
    style(ws1, f"A{row}", value=elem,   bold=True, bg=bg)
    style(ws1, f"B{row}", value=formula, bg=bg, align="left")
    style(ws1, f"C{row}", formula=a0,   bg=bg)
    style(ws1, f"D{row}", value="×",    bg=bg)
    style(ws1, f"E{row}", formula=b0,   bg=bg)
    style(ws1, f"F{row}", value="+",    bg=bg)
    style(ws1, f"G{row}", formula=a1,   bg=bg)
    style(ws1, f"H{row}", value="×",    bg=bg)
    style(ws1, f"I{row}", formula=b1,   bg=bg)
    style(ws1, f"J{row}", value="=",    bg=bg)
    style(ws1, f"K{row}", formula=sumf, bg=YELLOW, bold=True)
    style(ws1, f"L{row}", value=exp,    bg=GREEN,  bold=True)
    style(ws1, f"M{row}", formula=f'=IF(K{row}=L{row},"✓ PASS","✗ FAIL")',
          bg=GREEN, bold=True, fg=DARK_GREEN)
    set_row_height(ws1, row, 20)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PE Weights
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2 - PE Weights")
ws2.sheet_view.showGridLines = False

for col, w in enumerate([4, 18, 18, 18], 1):
    set_col_width(ws2, col, w)
for row in range(1, 30):
    set_row_height(ws2, row, 22)

ws2.merge_cells("A1:D1")
style(ws2, "A1", value="PE Weight Assignment:  PE[i][j] preloads B[i][j]",
      bold=True, bg=DARK_BLUE, fg=WHITE, size=13, align="center", border=False)
set_row_height(ws2, 1, 30)

# PE grid headers
merge_header(ws2, "A3:D3", "Preloaded Weight Grid", bg=MED_BLUE)
sub_header(ws2, "A4", "", bg=MED_BLUE)
sub_header(ws2, "B4", "Column 0  (n=0)", bg=MED_BLUE)
sub_header(ws2, "C4", "Column 1  (n=1)", bg=MED_BLUE)
sub_header(ws2, "D4", "Role in C output", bg=MED_BLUE)

for r, (row_label, pe00, pe01, desc00, desc01) in enumerate([
    ("Row 0\n(k=0)",
     "PE[0][0]\nweight = B[0][0] = 5",
     "PE[0][1]\nweight = B[0][1] = 6",
     "Partial product for C[*][0], k=0 term",
     "Partial product for C[*][1], k=0 term"),
    ("Row 1\n(k=1)",
     "PE[1][0]\nweight = B[1][0] = 7",
     "PE[1][1]\nweight = B[1][1] = 8",
     "Completes C[*][0] dot product",
     "Completes C[*][1] dot product"),
], 5):
    style(ws2, f"A{r}", value=row_label, bold=True, bg=LIGHT_BLUE, align="center")
    style(ws2, f"B{r}", value=pe00, bold=True, bg=YELLOW, size=12)
    style(ws2, f"C{r}", value=pe01, bold=True, bg=YELLOW, size=12)
    set_row_height(ws2, r, 36)

# Individual PE detail table
merge_header(ws2, "A8:D8", "Individual PE Detail", bg=MED_BLUE)
for col, h in enumerate(["PE", "Weight Source", "Weight Value", "Computes contribution to …"], 1):
    sub_header(ws2, f"{get_column_letter(col)}9", h, bg=LIGHT_BLUE, fg="000000")

detail = [
    ("PE[0][0]", "B[0][0]", 5, "C[0][0] and C[1][0]  (k=0 term for column 0)"),
    ("PE[0][1]", "B[0][1]", 6, "C[0][1] and C[1][1]  (k=0 term for column 1)"),
    ("PE[1][0]", "B[1][0]", 7, "C[0][0] and C[1][0]  (k=1 term for column 0)"),
    ("PE[1][1]", "B[1][1]", 8, "C[0][1] and C[1][1]  (k=1 term for column 1)"),
]
for i, (pe, src, val, role) in enumerate(detail, 10):
    bg = PALE_BLUE if i % 2 == 0 else WHITE
    style(ws2, f"A{i}", value=pe,   bold=True, bg=bg)
    style(ws2, f"B{i}", value=src,  bg=bg)
    style(ws2, f"C{i}", value=val,  bg=YELLOW, bold=True)
    style(ws2, f"D{i}", value=role, bg=bg, align="left")
    set_row_height(ws2, i, 22)

# Operation rule
merge_header(ws2, "A13:D13", "PE Operation Rule (every active cycle)", bg=DARK_BLUE)
ws2.merge_cells("A14:D14")
style(ws2, "A14",
      value="partial_sum_out  =  partial_sum_in  +  ( input  ×  weight )",
      bold=True, bg=YELLOW, size=12, align="center")
set_row_height(ws2, 14, 26)

ws2.merge_cells("A15:D15")
style(ws2, "A15",
      value="• partial_sum_in comes from the PE directly above (0 for Row 0)     "
            "• input is the A element broadcast into this row from the left",
      bg=PALE_BLUE, align="left", italic=True)
set_row_height(ws2, 15, 22)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Cycle-by-Cycle Trace
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3 - Cycle Trace")
ws3.sheet_view.showGridLines = False

for col, w in enumerate([6, 16, 16, 22, 22, 22, 22, 20], 1):
    set_col_width(ws3, col, w)
for row in range(1, 120):
    set_row_height(ws3, row, 20)

ws3.merge_cells("A1:H1")
style(ws3, "A1",
      value="Cycle-by-Cycle Dataflow Trace  |  C = A × B  |  Weight-Stationary",
      bold=True, bg=DARK_BLUE, fg=WHITE, size=13, align="center", border=False)
set_row_height(ws3, 1, 30)

# ── Summary table first ───────────────────────────────────────────────────────
merge_header(ws3, "A3:H3", "Summary Table", bg=MED_BLUE)
hdrs = ["Cycle", "Input → Row 0", "Input → Row 1",
        "PE[0][0]  partial sum", "PE[0][1]  partial sum",
        "PE[1][0]  partial sum", "PE[1][1]  partial sum",
        "Output C"]
for col, h in enumerate(hdrs, 1):
    sub_header(ws3, f"{get_column_letter(col)}4", h, bg=LIGHT_BLUE, fg="000000")
set_row_height(ws3, 4, 28)

summary = [
    (1, "A[0][0] = 1", "bubble (0)",
     "0+(1×5)=5", "0+(1×6)=6", "0+(0×7)=0", "0+(0×8)=0", "—"),
    (2, "bubble (0)", "A[0][1] = 2",
     "holds 5", "holds 6", "5+(2×7)=19", "6+(2×8)=22", "C[0]=[19, 22]"),
    (3, "A[1][0] = 3", "bubble (0)",
     "0+(3×5)=15", "0+(3×6)=18", "0+(0×7)=0", "0+(0×8)=0", "—"),
    (4, "bubble (0)", "A[1][1] = 4",
     "holds 15", "holds 18", "15+(4×7)=43", "18+(4×8)=50", "C[1]=[43, 50]"),
]
for i, row_data in enumerate(summary, 5):
    cyc, r0, r1, p00, p01, p10, p11, out = row_data
    bg = PALE_BLUE if i % 2 == 1 else WHITE
    style(ws3, f"A{i}", value=cyc,  bold=True, bg=bg)
    style(ws3, f"B{i}", value=r0,   bg=bg)
    style(ws3, f"C{i}", value=r1,   bg=bg)
    out_c = YELLOW if "holds" not in p00 and cyc in (1,3) else YELLOW
    style(ws3, f"D{i}", value=p00, bg=YELLOW if cyc in (1,3) else PALE_BLUE)
    style(ws3, f"E{i}", value=p01, bg=YELLOW if cyc in (1,3) else PALE_BLUE)
    style(ws3, f"F{i}", value=p10,
          bg=GREEN if "19" in p10 or "43" in p10 else PALE_BLUE,
          bold=("19" in p10 or "43" in p10))
    style(ws3, f"G{i}", value=p11,
          bg=GREEN if "22" in p11 or "50" in p11 else PALE_BLUE,
          bold=("22" in p11 or "50" in p11))
    style(ws3, f"H{i}", value=out, bold=(out != "—"),
          bg=GREEN if out != "—" else WHITE, fg=DARK_GREEN if out != "—" else "000000")
    set_row_height(ws3, i, 22)

# ── Per-cycle detail ──────────────────────────────────────────────────────────
cycle_data = [
    # (cycle_num, row0_input_label, row0_input_val,
    #  row1_input_label, row1_input_val,
    #  pe_details: list of (pe_name, ps_in, inp, wt, ps_out, note, bg),
    #  output_label)
    (1, "A[0][0]", 1, "bubble", 0,
     [
         ("PE[0][0]", 0, 1, 5, None, "→ 0 + (1 × 5) = 5", YELLOW),
         ("PE[0][1]", 0, 1, 6, None, "→ 0 + (1 × 6) = 6", YELLOW),
         ("PE[1][0]", 0, 0, 7, None, "→ 0 + (0 × 7) = 0  [bubble]", PALE_BLUE),
         ("PE[1][1]", 0, 0, 8, None, "→ 0 + (0 × 8) = 0  [bubble]", PALE_BLUE),
     ],
     "No output"),
    (2, "bubble", 0, "A[0][1]", 2,
     [
         ("PE[0][0]", 5, None, 5, None, "holds partial sum = 5 (drains to PE[1][0])", PALE_BLUE),
         ("PE[0][1]", 6, None, 6, None, "holds partial sum = 6 (drains to PE[1][1])", PALE_BLUE),
         ("PE[1][0]", 5, 2, 7, None, "→ 5 + (2 × 7) = 5 + 14 = 19  → C[0][0]", GREEN),
         ("PE[1][1]", 6, 2, 8, None, "→ 6 + (2 × 8) = 6 + 16 = 22  → C[0][1]", GREEN),
     ],
     "OUTPUT  C[0] = [19, 22]"),
    (3, "A[1][0]", 3, "bubble", 0,
     [
         ("PE[0][0]", 0, 3, 5, None, "→ 0 + (3 × 5) = 15  [row 0 partial sum reset]", YELLOW),
         ("PE[0][1]", 0, 3, 6, None, "→ 0 + (3 × 6) = 18  [row 0 partial sum reset]", YELLOW),
         ("PE[1][0]", 0, 0, 7, None, "→ 0 + (0 × 7) = 0   [bubble / reset]", PALE_BLUE),
         ("PE[1][1]", 0, 0, 8, None, "→ 0 + (0 × 8) = 0   [bubble / reset]", PALE_BLUE),
     ],
     "No output"),
    (4, "bubble", 0, "A[1][1]", 4,
     [
         ("PE[0][0]", 15, None, 5, None, "holds partial sum = 15 (drains to PE[1][0])", PALE_BLUE),
         ("PE[0][1]", 18, None, 6, None, "holds partial sum = 18 (drains to PE[1][1])", PALE_BLUE),
         ("PE[1][0]", 15, 4, 7, None, "→ 15 + (4 × 7) = 15 + 28 = 43  → C[1][0]", GREEN),
         ("PE[1][1]", 18, 4, 8, None, "→ 18 + (4 × 8) = 18 + 32 = 50  → C[1][1]", GREEN),
     ],
     "OUTPUT  C[1] = [43, 50]"),
]

# Detailed PE columns
detail_hdrs = ["PE", "ps_in (from above)", "input (from left)",
               "weight (fixed)", "Calculation", "ps_out / Result"]

cur_row = 11
for cyc_num, r0_lbl, r0_val, r1_lbl, r1_val, pes, out_lbl in cycle_data:
    # Cycle header
    ws3.merge_cells(f"A{cur_row}:H{cur_row}")
    cyc_bg = MED_BLUE if cyc_num % 2 == 1 else DARK_BLUE
    style(ws3, f"A{cur_row}",
          value=f"CYCLE {cyc_num}  |  Row 0 input: {r0_lbl} = {r0_val}"
                f"   |   Row 1 input: {r1_lbl} = {r1_val}",
          bold=True, bg=cyc_bg, fg=WHITE, size=12, align="left")
    set_row_height(ws3, cur_row, 26)
    cur_row += 1

    # Sub-headers for detail columns
    for col, h in enumerate(detail_hdrs, 1):
        sub_header(ws3, f"{get_column_letter(col)}{cur_row}", h,
                   bg=LIGHT_BLUE, fg="000000")
    # merge columns 5-6 as one "Calculation" and 7-8 as "ps_out"? No, keep 6 cols
    # Actually map: A=PE, B=ps_in, C=input, D=weight, E=Calculation, F=ps_out → use A-F, merge G-H
    ws3.merge_cells(f"G{cur_row}:H{cur_row}")
    style(ws3, f"G{cur_row}", value="Notes / Output", bold=True,
          bg=LIGHT_BLUE, fg="000000")
    set_row_height(ws3, cur_row, 22)
    cur_row += 1

    for pe_name, ps_in, inp, wt, ps_out_override, note, bg in pes:
        # Calculate ps_out
        if inp is None:
            calc = f"holds {ps_in}"
            ps_out = ps_in
        else:
            ps_out = ps_in + inp * wt
            calc = f"{ps_in} + ({inp} × {wt}) = {ps_in} + {inp*wt} = {ps_out}"

        style(ws3, f"A{cur_row}", value=pe_name,  bold=True, bg=bg)
        style(ws3, f"B{cur_row}", value=ps_in,     bg=bg)
        style(ws3, f"C{cur_row}", value=inp if inp is not None else "—", bg=bg)
        style(ws3, f"D{cur_row}", value=wt,        bg=bg)
        style(ws3, f"E{cur_row}", value=calc,       bg=YELLOW if bg == YELLOW else bg,
              align="left")
        style(ws3, f"F{cur_row}", value=ps_out,    bold=True,
              bg=GREEN if bg == GREEN else YELLOW if bg == YELLOW else bg)
        ws3.merge_cells(f"G{cur_row}:H{cur_row}")
        style(ws3, f"G{cur_row}", value=note, bg=bg, align="left",
              italic=True, fg=DARK_GREEN if bg == GREEN else DARK_GREY)
        set_row_height(ws3, cur_row, 22)
        cur_row += 1

    # Output row
    out_bg = GREEN if "OUTPUT" in out_lbl else GREY
    ws3.merge_cells(f"A{cur_row}:H{cur_row}")
    style(ws3, f"A{cur_row}", value=f"⟹  {out_lbl}",
          bold=True, bg=out_bg,
          fg=DARK_GREEN if "OUTPUT" in out_lbl else DARK_GREY,
          size=12, align="center")
    set_row_height(ws3, cur_row, 24)
    cur_row += 2  # blank gap

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Counts
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4 - Counts")
ws4.sheet_view.showGridLines = False

for col, w in enumerate([6, 14, 16, 16, 22, 14, 14], 1):
    set_col_width(ws4, col, w)
for row in range(1, 100):
    set_row_height(ws4, row, 20)

ws4.merge_cells("A1:G1")
style(ws4, "A1", value="CF05 Counts  —  MAC Operations · Input Reuse · Off-Chip Accesses",
      bold=True, bg=DARK_BLUE, fg=WHITE, size=13, align="center", border=False)
set_row_height(ws4, 1, 30)

# ── (a) MAC Operations ────────────────────────────────────────────────────────
merge_header(ws4, "A3:G3", "(a)  Total MAC Operations", bg=MED_BLUE)
for col, h in enumerate(
    ["#", "PE", "Cycle", "ps_in", "input × weight", "product", "ps_out  (=MAC result)"], 1):
    sub_header(ws4, f"{get_column_letter(col)}4", h, bg=LIGHT_BLUE, fg="000000")
set_row_height(ws4, 4, 22)

macs = [
    (1,  "PE[0][0]", 1,  0, "1 × 5", 5,  5),
    (2,  "PE[0][1]", 1,  0, "1 × 6", 6,  6),
    (3,  "PE[1][0]", 2,  5, "2 × 7", 14, 19),
    (4,  "PE[1][1]", 2,  6, "2 × 8", 16, 22),
    (5,  "PE[0][0]", 3,  0, "3 × 5", 15, 15),
    (6,  "PE[0][1]", 3,  0, "3 × 6", 18, 18),
    (7,  "PE[1][0]", 4, 15, "4 × 7", 28, 43),
    (8,  "PE[1][1]", 4, 18, "4 × 8", 32, 50),
]
for i, (num, pe, cyc, ps_in, mult, prod, ps_out) in enumerate(macs, 5):
    bg = PALE_BLUE if i % 2 == 1 else WHITE
    is_output = cyc in (2, 4) and "PE[1]" in pe
    style(ws4, f"A{i}", value=f"MAC #{num}", bold=True, bg=bg)
    style(ws4, f"B{i}", value=pe,            bg=bg)
    style(ws4, f"C{i}", value=cyc,           bg=bg)
    style(ws4, f"D{i}", value=ps_in,         bg=bg)
    style(ws4, f"E{i}", value=mult,          bg=YELLOW)
    style(ws4, f"F{i}", value=prod,          bg=YELLOW)
    style(ws4, f"G{i}", value=ps_out, bold=is_output,
          bg=GREEN if is_output else YELLOW,
          fg=DARK_GREEN if is_output else "000000")
    set_row_height(ws4, i, 20)

# Total row
merge_header(ws4, "A13:F13", "TOTAL MAC OPERATIONS  =  8", bg=DARK_BLUE)
style(ws4, "G13", formula="=COUNTA(A5:A12)", bold=True, bg=YELLOW, size=12)
ws4.merge_cells("A14:G14")
style(ws4, "A14",
      value="Cross-check:  M × K × N = 2 × 2 × 2 = 8  ✓   (every A×B element contributes exactly one MAC)",
      bg=GREEN, italic=True, align="left", fg=DARK_GREEN)
set_row_height(ws4, 14, 22)

# ── (b) Input Reuse ───────────────────────────────────────────────────────────
merge_header(ws4, "A17:G17", "(b)  Input Value Reuse", bg=MED_BLUE)
for col, h in enumerate(
    ["A element", "Value", "Cycle used", "PE use #1", "Computation #1",
     "PE use #2", "Computation #2"], 1):
    sub_header(ws4, f"{get_column_letter(col)}18", h, bg=LIGHT_BLUE, fg="000000")
set_row_height(ws4, 18, 22)

reuse = [
    ("A[0][0]", 1, 1, "PE[0][0]", "1 × 5 = 5",  "PE[0][1]", "1 × 6 = 6"),
    ("A[0][1]", 2, 2, "PE[1][0]", "2 × 7 = 14", "PE[1][1]", "2 × 8 = 16"),
    ("A[1][0]", 3, 3, "PE[0][0]", "3 × 5 = 15", "PE[0][1]", "3 × 6 = 18"),
    ("A[1][1]", 4, 4, "PE[1][0]", "4 × 7 = 28", "PE[1][1]", "4 × 8 = 32"),
]
for i, (elem, val, cyc, pe1, comp1, pe2, comp2) in enumerate(reuse, 19):
    bg = PALE_BLUE if i % 2 == 1 else WHITE
    style(ws4, f"A{i}", value=elem,  bold=True, bg=bg)
    style(ws4, f"B{i}", value=val,   bg=bg)
    style(ws4, f"C{i}", value=cyc,   bg=bg)
    style(ws4, f"D{i}", value=pe1,   bg=YELLOW)
    style(ws4, f"E{i}", value=comp1, bg=YELLOW)
    style(ws4, f"F{i}", value=pe2,   bg=ORANGE)
    style(ws4, f"G{i}", value=comp2, bg=ORANGE)
    set_row_height(ws4, i, 20)

merge_header(ws4, "A23:G23",
             "Each input value is fetched ONCE from off-chip and used by 2 PEs  →  reuse factor = 2×",
             bg=DARK_BLUE)
set_row_height(ws4, 23, 24)

# ── (c) Off-Chip Memory Accesses ──────────────────────────────────────────────
merge_header(ws4, "A26:G26", "(c)  Off-Chip Memory Accesses", bg=MED_BLUE)

# A reads
merge_header(ws4, "A27:G27", "A  —  Input reads  (fetched during streaming)", bg=LIGHT_BLUE, fg="000000")
for col, h in enumerate(["Element", "Value", "Fetched at cycle",
                          "PE(s) served", "Access type", "Count", ""], 1):
    sub_header(ws4, f"{get_column_letter(col)}28", h, bg=PALE_BLUE, fg="000000")

a_reads = [
    ("A[0][0]", 1, 1, "PE[0][0], PE[0][1]", "off-chip read"),
    ("A[0][1]", 2, 2, "PE[1][0], PE[1][1]", "off-chip read"),
    ("A[1][0]", 3, 3, "PE[0][0], PE[0][1]", "off-chip read"),
    ("A[1][1]", 4, 4, "PE[1][0], PE[1][1]", "off-chip read"),
]
for i, (elem, val, cyc, pes, acc) in enumerate(a_reads, 29):
    bg = PALE_BLUE if i % 2 == 1 else WHITE
    style(ws4, f"A{i}", value=elem, bold=True, bg=bg)
    style(ws4, f"B{i}", value=val,  bg=bg)
    style(ws4, f"C{i}", value=cyc,  bg=bg)
    style(ws4, f"D{i}", value=pes,  bg=bg)
    style(ws4, f"E{i}", value=acc,  bg=LIGHT_BLUE)
    style(ws4, f"F{i}", value=1,    bg=LIGHT_BLUE, bold=True)
    style(ws4, f"G{i}", value="",   bg=bg)

merge_header(ws4, "A33:E33", "A subtotal:", bg=LIGHT_BLUE, fg="000000")
style(ws4, "F33", formula="=SUM(F29:F32)", bold=True, bg=YELLOW, size=12)
style(ws4, "G33", value="reads", bg=YELLOW)

# B preloads
merge_header(ws4, "A35:G35", "B  —  Weight preloads  (loaded once before compute begins)", bg=LIGHT_BLUE, fg="000000")
for col, h in enumerate(["Element", "Value", "Loaded into PE",
                          "Re-fetched during compute?", "Access type", "Count", ""], 1):
    sub_header(ws4, f"{get_column_letter(col)}36", h, bg=PALE_BLUE, fg="000000")

b_reads = [
    ("B[0][0]", 5, "PE[0][0]", "No — stays fixed", "off-chip read (preload)"),
    ("B[0][1]", 6, "PE[0][1]", "No — stays fixed", "off-chip read (preload)"),
    ("B[1][0]", 7, "PE[1][0]", "No — stays fixed", "off-chip read (preload)"),
    ("B[1][1]", 8, "PE[1][1]", "No — stays fixed", "off-chip read (preload)"),
]
for i, (elem, val, pe, refetch, acc) in enumerate(b_reads, 37):
    bg = PALE_BLUE if i % 2 == 1 else WHITE
    style(ws4, f"A{i}", value=elem,    bold=True, bg=bg)
    style(ws4, f"B{i}", value=val,     bg=bg)
    style(ws4, f"C{i}", value=pe,      bg=bg)
    style(ws4, f"D{i}", value=refetch, bg=GREEN, fg=DARK_GREEN, italic=True)
    style(ws4, f"E{i}", value=acc,     bg=LIGHT_BLUE)
    style(ws4, f"F{i}", value=1,       bg=LIGHT_BLUE, bold=True)
    style(ws4, f"G{i}", value="",      bg=bg)

merge_header(ws4, "A41:E41", "B subtotal:", bg=LIGHT_BLUE, fg="000000")
style(ws4, "F41", formula="=SUM(F37:F40)", bold=True, bg=YELLOW, size=12)
style(ws4, "G41", value="reads", bg=YELLOW)

# C writes
merge_header(ws4, "A43:G43", "C  —  Output writes  (written when bottom row produces result)", bg=LIGHT_BLUE, fg="000000")
for col, h in enumerate(["Element", "Value", "Written at cycle",
                          "Produced by", "Access type", "Count", ""], 1):
    sub_header(ws4, f"{get_column_letter(col)}44", h, bg=PALE_BLUE, fg="000000")

c_writes = [
    ("C[0][0]", 19, 2, "PE[1][0]", "off-chip write"),
    ("C[0][1]", 22, 2, "PE[1][1]", "off-chip write"),
    ("C[1][0]", 43, 4, "PE[1][0]", "off-chip write"),
    ("C[1][1]", 50, 4, "PE[1][1]", "off-chip write"),
]
for i, (elem, val, cyc, pe, acc) in enumerate(c_writes, 45):
    bg = PALE_BLUE if i % 2 == 1 else WHITE
    style(ws4, f"A{i}", value=elem, bold=True, bg=bg)
    style(ws4, f"B{i}", value=val,  bold=True, bg=GREEN, fg=DARK_GREEN)
    style(ws4, f"C{i}", value=cyc,  bg=bg)
    style(ws4, f"D{i}", value=pe,   bg=bg)
    style(ws4, f"E{i}", value=acc,  bg=LIGHT_BLUE)
    style(ws4, f"F{i}", value=1,    bg=LIGHT_BLUE, bold=True)
    style(ws4, f"G{i}", value="",   bg=bg)

merge_header(ws4, "A49:E49", "C subtotal:", bg=LIGHT_BLUE, fg="000000")
style(ws4, "F49", formula="=SUM(F45:F48)", bold=True, bg=YELLOW, size=12)
style(ws4, "G49", value="writes", bg=YELLOW)

# Grand total
merge_header(ws4, "A51:E51", "GRAND TOTAL off-chip accesses:", bg=DARK_BLUE)
style(ws4, "F51", formula="=F33+F41+F49", bold=True, bg=GREEN, size=14, fg=DARK_GREEN)
style(ws4, "G51", value="total", bold=True, bg=GREEN, fg=DARK_GREEN)
set_row_height(ws4, 51, 28)

ws4.merge_cells("A52:G52")
style(ws4, "A52",
      value="= 4 (A reads) + 4 (B preloads) + 4 (C writes) = 12  "
            "— B never re-fetched during compute (weight-stationary advantage)",
      bg=PALE_BLUE, italic=True, align="left", fg=DARK_GREY)
set_row_height(ws4, 52, 22)

# ── (d) Output-Stationary ────────────────────────────────────────────────────
merge_header(ws4, "A55:G55", "(d)  Output-Stationary: What would stay fixed?", bg=MED_BLUE)
ws4.merge_cells("A56:G56")
style(ws4, "A56",
      value="In output-stationary dataflow, the partial sum for each output element C[i][j] "
            "stays fixed inside its dedicated PE, while both the A row elements and B column "
            "elements stream through — so the accumulated output values (partial sums) would "
            "be stationary rather than the weights.",
      bg=PALE_BLUE, align="left", italic=True)
set_row_height(ws4, 56, 50)

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = "/home/bao/ECE410-HW4AI/codefest/cf05/systolic_trace.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
